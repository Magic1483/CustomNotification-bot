import openpyxl
import datetime

def sheetpiece_out(n,m,EXCel_file):
    wb_obj = openpyxl.load_workbook(EXCel_file)
    
    sheet_obj =wb_obj.active

    max_row = sheet_obj.max_row
    max_col = sheet_obj.max_column

    if m=='max':
        m=max_row

    # Will print a particular row value
    def __output_row(nrow,ncol):
        cell_obj = sheet_obj.cell(nrow, ncol)
        return cell_obj.value


    #print(output_row(2,1))

    def __full_row(row):
        tmp=''
        for i in range(1,3):
            tmp=tmp+' '+__output_row(row,i)
        return tmp


    #20-25
    def __sheetpiece_out_asist(n,m):   
        tmp='' 
        for i in range(n,m+1):
            tmp=__full_row(i)+'\n'+tmp
        return tmp

    return __sheetpiece_out_asist(n,m)


def input_note(name ,link):
    wb = openpyxl.load_workbook("Notes.xlsx")
    worksheet = wb['notes'] 

    rows = worksheet.max_row
    global row
    row=rows+1

    def __input_notes(name,link): 
        global row
        worksheet[f'A{row}']=name 
        worksheet[f'B{row}']=link 
        wb.save('Notes.xlsx') 
        rows = worksheet.max_row
    
    __input_notes(name,link)
    
